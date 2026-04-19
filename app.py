from flask import (
    Flask, render_template, request, redirect,
    url_for, send_file, flash, jsonify, abort, session
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os as _os
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo

_BRT = ZoneInfo("America/Sao_Paulo")
from io import BytesIO
import json
import platform
import re
import subprocess
import unicodedata
import uuid

from gerar_contrato import gerar_docx, gerar_termo_quitacao, gerar_notificacao_avalista, gerar_notificacao_inadimplente, gerar_vistoria_entrega, gerar_vistoria_nova, nome_arquivo_saida

app = Flask(__name__)
app.secret_key = _os.environ.get("SECRET_KEY", "ativuz-secret-dev-2026")


@app.errorhandler(Exception)
def handle_any_error(e):
    import traceback; traceback.print_exc()
    return jsonify({"error": str(e)}), 500


# ── Autenticação ──────────────────────────────────────────────────────────────

_ROTAS_PUBLICAS = {"login", "static", "admin_novo_usuario"}

@app.before_request
def verificar_login():
    if request.endpoint in _ROTAS_PUBLICAS:
        return
    if not session.get("usuario"):
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("usuario"):
        return redirect(url_for("dashboard"))
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        sb = _supabase()
        if not sb:
            erro = "Serviço indisponível. Tente novamente."
        else:
            try:
                res = sb.table("usuarios").select("nome, senha_hash") \
                    .eq("nome", nome).eq("ativo", True).single().execute()
                usuario = res.data
                if usuario and check_password_hash(usuario["senha_hash"], senha):
                    session["usuario"] = nome
                    return redirect(url_for("dashboard"))
                else:
                    erro = "Nome ou senha incorretos."
            except Exception:
                erro = "Nome ou senha incorretos."
    return render_template("login.html", erro=erro)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin/novo-usuario", methods=["GET", "POST"])
def admin_novo_usuario():
    token_correto = _os.environ.get("ADMIN_TOKEN", "")
    token = request.args.get("token", "")
    if not token_correto or token != token_correto:
        abort(403)
    mensagem = None
    erro = None
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        if not nome or not senha:
            erro = "Nome e senha são obrigatórios."
        else:
            sb = _supabase()
            if not sb:
                erro = "Supabase não configurado."
            else:
                try:
                    sb.table("usuarios").insert({
                        "nome": nome,
                        "senha_hash": generate_password_hash(senha),
                        "ativo": True,
                    }).execute()
                    mensagem = f"Usuário '{nome}' criado com sucesso!"
                except Exception as exc:
                    erro = f"Erro ao criar usuário: {exc}"
    return f"""
    <!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
    <title>Novo Usuário — Admin</title>
    <style>body{{font-family:Inter,sans-serif;background:#f0f2f7;display:flex;
    align-items:center;justify-content:center;min-height:100vh;}}
    .card{{background:#fff;border-radius:14px;padding:2rem;width:360px;
    box-shadow:0 4px 20px rgba(0,0,0,.1);}}
    h1{{font-size:1.1rem;margin-bottom:1.5rem;}}
    label{{font-size:.75rem;font-weight:600;text-transform:uppercase;
    letter-spacing:.05em;color:#475569;display:block;margin-bottom:.3rem;}}
    input{{width:100%;padding:.6rem .8rem;border:1.5px solid #e2e8f0;border-radius:8px;
    font-family:inherit;font-size:.9rem;margin-bottom:1rem;outline:none;}}
    input:focus{{border-color:#4361ee;}}
    button{{width:100%;padding:.7rem;background:#4361ee;color:#fff;border:none;
    border-radius:8px;font-weight:700;font-size:.9rem;cursor:pointer;}}
    .ok{{color:#166534;background:#f0fdf4;border:1px solid #bbf7d0;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    .err{{color:#991b1b;background:#fef2f2;border:1px solid #fecaca;
    border-radius:8px;padding:.6rem .9rem;font-size:.85rem;margin-bottom:1rem;}}
    </style></head><body><div class="card">
    <h1>Criar novo usuário</h1>
    {"<div class='ok'>"+mensagem+"</div>" if mensagem else ""}
    {"<div class='err'>"+erro+"</div>" if erro else ""}
    <form method="POST" action="?token={token}">
    <label>Nome</label><input name="nome" required>
    <label>Senha</label><input type="password" name="senha" required>
    <button>Criar usuário</button></form></div></body></html>
    """

# ── Supabase (opcional — só ativa se as env vars estiverem definidas) ─────────

_sb = None

def _supabase():
    global _sb
    if _sb is None:
        url = _os.environ.get("SUPABASE_URL", "")
        key = _os.environ.get("SUPABASE_KEY", "")
        if url and key:
            from supabase import create_client
            _sb = create_client(url, key)
    return _sb

UPLOAD_FOLDER = Path("uploads")
CONTRATOS_FOLDER = Path("contratos")
TEMP_FOLDER = Path("temp_preview")
DOCX_TEMPLATES = Path("docx_templates")

UPLOAD_FOLDER.mkdir(exist_ok=True)
CONTRATOS_FOLDER.mkdir(exist_ok=True)
TEMP_FOLDER.mkdir(exist_ok=True)
DOCX_TEMPLATES.mkdir(exist_ok=True)


# ── helpers ───────────────────────────────────────────────

def _converter_pdf(caminho_docx: str, caminho_pdf: str):
    """Converte .docx para PDF: Word no Windows, LibreOffice no Linux."""
    if platform.system() == "Windows":
        import pythoncom
        from docx2pdf import convert
        pythoncom.CoInitialize()
        try:
            convert(caminho_docx, caminho_pdf)
        finally:
            pythoncom.CoUninitialize()
    else:
        output_dir = str(Path(caminho_pdf).parent)
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", output_dir, caminho_docx],
            check=True, capture_output=True,
        )
        # LibreOffice nomeia o PDF pelo stem do docx
        gerado = Path(output_dir) / (Path(caminho_docx).stem + ".pdf")
        if gerado != Path(caminho_pdf):
            gerado.rename(caminho_pdf)


def _slugify(texto: str) -> str:
    """Maiúsculas, sem acentos, espaços → underscore, sem caracteres especiais."""
    norm = unicodedata.normalize('NFD', texto)
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    norm = norm.upper().strip()
    norm = re.sub(r'[^A-Z0-9\s]', '', norm)
    norm = re.sub(r'\s+', '_', norm)
    return norm or "SEM_NOME"


def detectar_tipo(filename: str):
    """Retorna o tipo do template com base no nome do arquivo."""
    norm = unicodedata.normalize('NFD', filename.lower())
    norm = ''.join(c for c in norm if unicodedata.category(c) != 'Mn')
    if 'quitacao' in norm:
        return 'quitacao'
    if 'locacao' in norm:
        return 'locacao'
    if 'notificacao' in norm and 'inadimplente' in norm:
        return 'inadimplente'
    if 'notificacao' in norm:
        return 'notificacao'
    return None


def get_templates():
    result = []
    files = sorted(
        list(UPLOAD_FOLDER.glob("*.docx")) + list(UPLOAD_FOLDER.glob("*.xlsx")),
        key=lambda f: f.name,
    )
    for f in files:
        meta_path = UPLOAD_FOLDER / f"{f.stem}.json"
        display_name = f.stem
        if meta_path.exists():
            display_name = json.loads(meta_path.read_text(encoding="utf-8")).get("nome", f.stem)
        result.append({
            "filename": f.name,
            "nome": display_name,
            "tamanho_kb": round(f.stat().st_size / 1024, 1),
            "data": datetime.fromtimestamp(f.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        })
    return result


def _historico_append(locatario_nome: str, template: str, arquivo: str):
    sb = _supabase()
    if not sb:
        return
    try:
        sb.table("historico_docs").insert({
            "locatario_nome": locatario_nome,
            "template": template,
            "arquivo": arquivo,
            "data_hora": datetime.now(_BRT).strftime("%d/%m/%Y %H:%M"),
        }).execute()
    except Exception:
        import traceback; traceback.print_exc()


def _gerar_para_caminho(form, tipo, template_path_str, caminho_saida):
    """Gera o documento para caminho_saida. Retorna nome_pessoa."""
    if tipo == "locacao":
        campos = [
            "locatario_nome", "locatario_rg", "locatario_cpf",
            "locatario_endereco", "locatario_cep", "locatario_telefone",
            "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
            "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
            "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
            "contrato_inicio", "contrato_duracao", "valor_semanal",
            "data_dia", "data_mes", "data_ano",
            "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
            "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
        ]
        dados = {c: form.get(c, "") for c in campos}
        gerar_docx(dados, caminho_saida, template_path=template_path_str)
        return dados["locatario_nome"]

    elif tipo == "notificacao":
        avalista_nome = form.get("avalista_nome_notif", "")
        gerar_notificacao_avalista(
            avalista_nome  = avalista_nome,
            data_contrato  = form.get("data_contrato", ""),
            locatario_nome = form.get("locatario_nome_notif", ""),
            valor_debito   = float(form.get("valor_debito") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
        )
        return avalista_nome

    elif tipo == "inadimplente":
        locatario_nome_inad = form.get("locatario_nome_inad", "")
        gerar_notificacao_inadimplente(
            locatario_nome = locatario_nome_inad,
            data_contrato  = form.get("data_contrato_inad", ""),
            valor_debito   = float(form.get("valor_debito_inad") or 0),
            caminho_saida  = caminho_saida,
            template_path  = template_path_str,
        )
        return locatario_nome_inad

    else:  # quitacao
        def _f(campo): return float(form.get(campo) or 0)
        def _i(campo): return int(float(form.get(campo) or 0))
        devedor_nome = form.get("devedor_nome", "")
        gerar_termo_quitacao(
            devedor_nome          = devedor_nome,
            devedor_cpf           = form.get("devedor_cpf", ""),
            placa                 = form.get("placa", ""),
            mes_referencia_fipe   = form.get("mes_referencia_fipe", ""),
            valor_fipe            = _f("valor_fipe"),
            percentual_fipe       = _f("percentual_fipe"),
            meias_diarias         = _f("meias_diarias"),
            entrada               = _f("entrada"),
            num_parcelas_pagas    = _i("num_parcelas_pagas"),
            valor_parcela_paga    = _f("valor_parcela_paga"),
            num_parcelas_semanais = _i("num_parcelas_semanais"),
            valor_parcela_semanal = _f("valor_parcela_semanal"),
            data_primeira_parcela = form.get("data_primeira_parcela", ""),
            data_assinatura       = form.get("data_assinatura", ""),
            caminho_saida         = caminho_saida,
            template_path         = template_path_str,
        )
        return devedor_nome


# ── página 1 — Templates ──────────────────────────────────

@app.route("/")
def dashboard():
    sb = _supabase()
    total_contratos = 0
    total_vistorias = 0
    total_docs = 0
    valor_mensal = "—"
    contratos = []
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, veiculo_placa, veiculo_marca, veiculo_modelo, contrato_inicio, valor_semanal",
                count="exact"
            ).order("criado_em", desc=True).limit(5).execute()
            contratos = res.data or []
            total_contratos = res.count or len(contratos)
        except Exception:
            pass
        try:
            rv = sb.table("vistorias").select("id", count="exact").execute()
            total_vistorias = rv.count or 0
        except Exception:
            pass
    if sb:
        try:
            rd = sb.table("historico_docs").select("id", count="exact").eq("deletado", False).execute()
            total_docs = rd.count or 0
        except Exception:
            pass
    return render_template(
        "dashboard.html",
        active="dashboard",
        total_contratos=total_contratos,
        total_vistorias=total_vistorias,
        total_docs=total_docs,
        valor_mensal=valor_mensal,
        contratos=contratos,
    )


@app.route("/templates")
def pagina_templates():
    return render_template("templates.html", templates=get_templates(), active="templates")


@app.route("/upload", methods=["POST"])
def upload_template():
    nome = request.form.get("nome", "").strip()
    arquivo = request.files.get("arquivo")

    if not nome:
        flash("Informe um nome para o template.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo .docx.", "erro")
        return redirect(url_for("pagina_templates"))

    if not arquivo.filename.lower().endswith((".docx", ".xlsx")):
        flash("Apenas arquivos .docx ou .xlsx são aceitos.", "erro")
        return redirect(url_for("pagina_templates"))

    uid = uuid.uuid4().hex[:8]
    safe_stem = secure_filename(f"{nome}_{uid}")
    dest = UPLOAD_FOLDER / f"{safe_stem}.docx"
    arquivo.save(str(dest))

    meta = UPLOAD_FOLDER / f"{safe_stem}.json"
    meta.write_text(json.dumps({"nome": nome}, ensure_ascii=False), encoding="utf-8")

    flash(f'Template "{nome}" enviado com sucesso!', "ok")
    return redirect(url_for("pagina_templates"))


@app.route("/templates/excluir/<filename>", methods=["POST"])
def excluir_template(filename):
    safe = secure_filename(filename)
    docx = UPLOAD_FOLDER / safe
    meta = UPLOAD_FOLDER / f"{Path(safe).stem}.json"
    if docx.exists():
        docx.unlink()
    if meta.exists():
        meta.unlink()
    flash("Template excluído.", "ok")
    return redirect(url_for("pagina_templates"))


# ── página 2 — Gerar Contrato ─────────────────────────────

@app.route("/gerar")
def pagina_gerar():
    return render_template("gerar.html", templates=get_templates(), active="gerar")


@app.route("/gerar-contrato", methods=["POST"])
def gerar_contrato_route():
    template_filename = request.form.get("template", "")
    if not template_filename:
        flash("Selecione um template.", "erro")
        return redirect(url_for("pagina_gerar"))

    template_path = UPLOAD_FOLDER / secure_filename(template_filename)
    if not template_path.exists():
        flash("Template não encontrado.", "erro")
        return redirect(url_for("pagina_gerar"))

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({
            "error": "Template não reconhecido. Renomeie o arquivo com 'locacao', 'quitacao', 'notificacao' ou 'inadimplente' no nome."
        }), 400

    formato = request.form.get("formato", "docx")

    # ── Nome do arquivo de saída ──────────────────────────
    if tipo == "locacao":
        ano        = datetime.now().strftime("%Y")
        nome_saida = f"{ano}_{_slugify(request.form.get('veiculo_placa', ''))}_{_slugify(request.form.get('locatario_nome', ''))}.docx"
    elif tipo == "notificacao":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_AVALISTA_{_slugify(request.form.get('avalista_nome_notif', ''))}_{data_slug}.docx"
    elif tipo == "inadimplente":
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"NOTIFICACAO_INADIMPLENTE_{_slugify(request.form.get('locatario_nome_inad', ''))}_{data_slug}.docx"
    else:  # quitacao
        data_slug  = datetime.now().strftime("%d.%m.%Y")
        nome_saida = f"QUITACAO_DIVIDA_{_slugify(request.form.get('devedor_nome', ''))}_{data_slug}.docx"

    caminho_saida = str(CONTRATOS_FOLDER / nome_saida)

    # ── Gerar documento ───────────────────────────────────
    try:
        nome_pessoa = _gerar_para_caminho(request.form, tipo, str(template_path), caminho_saida)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500

    # ── Histórico local ───────────────────────────────────
    meta_path = UPLOAD_FOLDER / f"{template_path.stem}.json"
    nome_template = template_filename
    if meta_path.exists():
        try:
            nome_template = json.loads(meta_path.read_text(encoding="utf-8")).get("nome", template_filename)
        except Exception:
            pass
    try:
        _historico_append(nome_pessoa, nome_template, nome_saida)
    except Exception:
        pass

    # ── Download direto ────────────────────────────────────
    if formato == "pdf":
        nome_pdf    = nome_saida.replace(".docx", ".pdf")
        caminho_pdf = str(CONTRATOS_FOLDER / nome_pdf)
        try:
            _converter_pdf(caminho_saida, caminho_pdf)
            pdf_bytes = BytesIO(Path(caminho_pdf).read_bytes())
            return send_file(
                pdf_bytes,
                as_attachment=True,
                download_name=nome_pdf,
                mimetype="application/pdf",
            )
        except Exception as e:
            docx_url = url_for("download_contrato", filename=nome_saida)
            return jsonify({
                "error": f"Erro ao gerar PDF: {e}",
                "docx_url": docx_url,
            }), 422

    return send_file(
        caminho_saida,
        as_attachment=True,
        download_name=nome_saida,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/preview-contrato", methods=["POST"])
def preview_contrato():
    import mammoth

    template_filename = request.form.get("template", "")
    if not template_filename:
        return jsonify({"error": "Selecione um template."}), 400

    template_path = UPLOAD_FOLDER / secure_filename(template_filename)
    if not template_path.exists():
        return jsonify({"error": "Template não encontrado."}), 400

    tipo = detectar_tipo(template_filename)
    if tipo is None:
        return jsonify({"error": "Template não reconhecido."}), 400
    if tipo == "vistoria":
        return jsonify({"error": "Pré-visualização não disponível para vistoria (formato .xlsx)."}), 400

    temp_id = uuid.uuid4().hex
    caminho_temp = str(TEMP_FOLDER / f"{temp_id}.docx")

    try:
        _gerar_para_caminho(request.form, tipo, str(template_path), caminho_temp)
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar pré-visualização: {e}"}), 500

    try:
        with open(caminho_temp, "rb") as f:
            result = mammoth.convert_to_html(f)
        html = result.value
    except Exception as e:
        Path(caminho_temp).unlink(missing_ok=True)
        return jsonify({"error": f"Erro ao converter para HTML: {e}"}), 500

    return jsonify({"html": html, "temp_id": temp_id})


@app.route("/cleanup-temp/<temp_id>", methods=["POST"])
def cleanup_temp(temp_id):
    if not re.match(r'^[0-9a-f]{32}$', temp_id):
        abort(400)
    caminho = TEMP_FOLDER / f"{temp_id}.docx"
    if caminho.exists():
        caminho.unlink()
    return jsonify({"ok": True})


# ── página 3 — Histórico ──────────────────────────────────

@app.route("/historico")
def pagina_historico():
    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    return render_template("historico.html", historico=historico, active="historico")


@app.route("/historico/download/<path:filename>")
def download_contrato(filename):
    caminho = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho.exists():
        flash("Arquivo não encontrado.", "erro")
        return redirect(url_for("pagina_historico"))
    ext = Path(filename).suffix.lower()
    mime = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if ext == ".xlsx"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    return send_file(
        str(caminho),
        as_attachment=True,
        download_name=Path(filename).name,
        mimetype=mime,
    )


@app.route("/historico/download-pdf/<path:filename>")
def download_contrato_pdf(filename):
    caminho_docx = (CONTRATOS_FOLDER / filename).resolve()
    if not str(caminho_docx).startswith(str(CONTRATOS_FOLDER.resolve())):
        abort(400)
    if not caminho_docx.exists():
        return jsonify({"error": "Arquivo não encontrado."}), 404

    nome_pdf    = Path(filename).stem + ".pdf"
    caminho_pdf = CONTRATOS_FOLDER / nome_pdf
    try:
        _converter_pdf(str(caminho_docx), str(caminho_pdf))
        pdf_bytes = BytesIO(Path(caminho_pdf).read_bytes())
        return send_file(
            pdf_bytes,
            as_attachment=True,
            download_name=nome_pdf,
            mimetype="application/pdf",
        )
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422


@app.route("/historico/excluir/<entry_id>", methods=["POST"])
def excluir_contrato(entry_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("historico_docs").update({"deletado": True}).eq("id", entry_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/exportar-excel")
def exportar_historico_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    sb = _supabase()
    historico = []
    if sb:
        try:
            res = sb.table("historico_docs").select("*") \
                .eq("deletado", False) \
                .order("criado_em", desc=True).execute()
            historico = res.data or []
        except Exception:
            pass
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    cabecalho = ["Locatário", "Template", "Data / Hora", "Nome do Arquivo"]
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for item in historico:
        ws.append([
            item.get("locatario_nome", ""),
            item.get("template", ""),
            item.get("data_hora", ""),
            item.get("arquivo", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    data_hoje = datetime.now().strftime("%d-%m-%Y")
    nome_arquivo = f"HISTORICO_ATIVUZ_{data_hoje}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Contrato de Locação — helpers e rotas ────────────────────────────────────

CONTRATO_LOCACAO_TEMPLATE = DOCX_TEMPLATES / "CONTRATO DE LOCAÇÃO EDITADO.docx"


def _salvar_contrato_locacao(insert: dict, caminho_docx: str, storage_path: str, edit_id: str = None):
    """INSERT no Supabase (main thread) + upload do arquivo (background)."""
    import threading, traceback as _tb
    sb = _supabase()
    if not sb:
        return
    try:
        sb.table("contratos_locacao").insert(insert).execute()
    except Exception:
        _tb.print_exc()
    if edit_id:
        try:
            old = sb.table("contratos_locacao").select("arquivo_path").eq("id", edit_id).single().execute()
            _old_path = (old.data or {}).get("arquivo_path")
            sb.table("contratos_locacao").delete().eq("id", edit_id).execute()
        except Exception:
            _old_path = None
            _tb.print_exc()
    else:
        _old_path = None

    _docx_bytes = Path(caminho_docx).read_bytes()
    _sp = storage_path
    _op = _old_path

    def _bg():
        try:
            sb2 = _supabase()
            if not sb2:
                return
            try:
                sb2.storage.from_("documentos").upload(
                    _sp, _docx_bytes,
                    {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                     "upsert": "true"},
                )
            except Exception:
                _tb.print_exc()
            if _op and _op != _sp:
                try:
                    sb2.storage.from_("documentos").remove([_op])
                except Exception:
                    pass
        except Exception:
            _tb.print_exc()

    threading.Thread(target=_bg, daemon=True).start()

_MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]


@app.route("/contrato-locacao")
def pagina_contrato_locacao():
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": agora.strftime("%d"),
        "data_mes": _MESES_PT[agora.month - 1],
        "data_ano": agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", defaults=defaults, active="contrato_locacao")


@app.route("/contrato-locacao/gerar", methods=["POST"])
def gerar_contrato_locacao_route():
    """Usado pelo fluxo de edição a partir de historico_contratos."""
    if not CONTRATO_LOCACAO_TEMPLATE.exists():
        return jsonify({"error": "Template não encontrado em docx_templates/."}), 404

    campos = [
        "locatario_nome", "locatario_rg", "locatario_cpf",
        "locatario_endereco", "locatario_cep", "locatario_telefone",
        "avalista_nome", "avalista_cpf", "avalista_endereco", "avalista_telefone",
        "veiculo_descricao", "veiculo_marca", "veiculo_modelo", "veiculo_ano",
        "veiculo_motor", "veiculo_chassi", "veiculo_cor", "veiculo_placa",
        "contrato_inicio", "contrato_duracao", "valor_semanal",
        "data_dia", "data_mes", "data_ano",
        "testemunha1_nome", "testemunha1_rg", "testemunha1_cpf",
        "testemunha2_nome", "testemunha2_rg", "testemunha2_cpf",
    ]
    dados   = {c: request.form.get(c, "") for c in campos}
    edit_id = request.form.get("edit_id", "").strip()

    placa_slug    = _slugify(dados.get("veiculo_placa") or "PLACA")
    nome_slug     = _slugify((dados.get("locatario_nome") or "LOCATARIO").split()[0])
    data_slug     = datetime.now(_BRT).strftime("%d.%m.%Y")
    nome_docx     = f"CONTRATO_LOCACAO_{placa_slug}_{nome_slug}_{data_slug}.docx"
    caminho_saida = str(CONTRATOS_FOLDER / nome_docx)

    try:
        gerar_docx(dados, caminho_saida, template_path=str(CONTRATO_LOCACAO_TEMPLATE))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao gerar contrato: {e}"}), 500

    _storage_path = f"contratos/{nome_docx}"
    _insert = {**dados, "arquivo_path": _storage_path}
    _salvar_contrato_locacao(_insert, caminho_saida, _storage_path, edit_id=edit_id or None)

    return jsonify({"redirect_url": url_for("historico_contratos")})


@app.route("/historico/contratos")
def historico_contratos():
    sb = _supabase()
    contratos = []
    erro = None
    if sb:
        try:
            res = sb.table("contratos_locacao").select(
                "id, locatario_nome, locatario_cpf, veiculo_placa, veiculo_marca, "
                "veiculo_modelo, contrato_inicio, valor_semanal, arquivo_path, criado_em"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            contratos = res.data or []
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado."
    return render_template("historico_contratos.html", contratos=contratos, erro=erro,
                           active="hist_contratos")


@app.route("/historico/contratos/<contrato_id>/excluir", methods=["POST"])
def excluir_contrato_locacao(contrato_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("contratos_locacao").update({"deletado": True}).eq("id", contrato_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/contratos/download/<contrato_id>")
def download_contrato_locacao_docx(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        path = res.data["arquivo_path"]
        signed = sb.storage.from_("documentos").create_signed_url(path, 60)
        return redirect(signed["signedURL"])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/historico/contratos/download/<contrato_id>/pdf")
def download_contrato_locacao_pdf(contrato_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("contratos_locacao").select("arquivo_path").eq("id", contrato_id).single().execute()
        docx_path = res.data["arquivo_path"]
        docx_bytes = sb.storage.from_("documentos").download(docx_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    tmp_docx = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    tmp_pdf  = tmp_docx.with_suffix(".pdf")
    try:
        tmp_docx.write_bytes(docx_bytes)
        _converter_pdf(str(tmp_docx), str(tmp_pdf))
        pdf_bytes = tmp_pdf.read_bytes()
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422
    finally:
        tmp_docx.unlink(missing_ok=True)
        tmp_pdf.unlink(missing_ok=True)

    nome_pdf = Path(docx_path).stem + ".pdf"
    return send_file(BytesIO(pdf_bytes), as_attachment=True,
                     download_name=nome_pdf, mimetype="application/pdf")


@app.route("/historico/contratos/<contrato_id>/editar")
def editar_contrato_locacao(contrato_id):
    sb = _supabase()
    if not sb:
        flash("Supabase não configurado.", "erro")
        return redirect(url_for("historico_contratos"))
    try:
        res = sb.table("contratos_locacao").select("*").eq("id", contrato_id).single().execute()
        contrato = res.data
    except Exception as e:
        flash(f"Erro ao buscar contrato: {e}", "erro")
        return redirect(url_for("historico_contratos"))
    agora = datetime.now(_BRT)
    defaults = {
        "data_dia": contrato.get("data_dia") or agora.strftime("%d"),
        "data_mes": contrato.get("data_mes") or _MESES_PT[agora.month - 1],
        "data_ano": contrato.get("data_ano") or agora.strftime("%Y"),
    }
    return render_template("contrato_locacao.html", contrato=contrato,
                           edit_id=contrato_id, defaults=defaults,
                           active="hist_contratos")


# ── Vistoria de Entrega ───────────────────────────────────────────────────────

VISTORIA_TEMPLATE = DOCX_TEMPLATES / "VISTORIA_TESTE_1.docx"


@app.route("/vistoria", methods=["GET"])
def pagina_vistoria():
    return render_template("vistoria.html", active="vistoria", vistoria=None, edit_id=None, acessorios={})


@app.route("/vistoria", methods=["POST"])
def processar_vistoria():
    if not VISTORIA_TEMPLATE.exists():
        flash("Template de vistoria não encontrado em docx_templates/.", "erro")
        return redirect(url_for("pagina_vistoria"))

    dados = {
        "cliente":                  request.form.get("cliente", ""),
        "tel":                      request.form.get("tel", ""),
        "preenchido_por":           request.form.get("preenchido_por", ""),
        "endereco":                 request.form.get("endereco", ""),
        "chassi":                   request.form.get("chassi", ""),
        "motor":                    request.form.get("motor", ""),
        "veiculo":                  request.form.get("veiculo", ""),
        "placa":                    request.form.get("placa", ""),
        "ano":                      request.form.get("ano", ""),
        "cor":                      request.form.get("cor", ""),
        "hodometro_entrega":        request.form.get("hodometro_entrega", ""),
        "hodometro_retorno":        request.form.get("hodometro_retorno", ""),
        "combustivel":              request.form.get("combustivel", ""),
        "data":                     datetime.now().strftime("%d/%m/%Y"),
        "danos":                    request.form.get("danos", ""),
        "observacoes":              request.form.get("observacoes", ""),
        "sintomas":                 request.form.get("sintomas", ""),
        "assinatura_cliente":       request.form.get("assinatura_cliente", ""),
        "assinatura_responsavel":   request.form.get("assinatura_responsavel", ""),
        # Acessórios (form fields sem prefixo, mapeados para chaves que gerar_contrato usa)
        "acc_calotas":          request.form.get("calotas", ""),
        "acc_buzina":           request.form.get("buzina", ""),
        "acc_doc_crlv":         request.form.get("doc_crlv", ""),
        "acc_triangulo":        request.form.get("triangulo", ""),
        "acc_antena":           request.form.get("antena", ""),
        "acc_sensor_re":        request.form.get("sensor_re", ""),
        "acc_som":              request.form.get("som", ""),
        "acc_tapetes":          request.form.get("tapetes", ""),
        "acc_limpadores":       request.form.get("limpadores", ""),
        "acc_chave_roda":       request.form.get("chave_roda", ""),
        "acc_vidros":           request.form.get("vidros", ""),
        "acc_oleo_motor":       request.form.get("oleo_motor", ""),
        "acc_alarme":           request.form.get("alarme", ""),
        "acc_lampadas":         request.form.get("lampadas", ""),
        "acc_macaco":           request.form.get("macaco", ""),
        "acc_estepe":           request.form.get("estepe", ""),
        "acc_gnv":              request.form.get("gnv", ""),
        "acc_agua":             request.form.get("agua", ""),
        "acc_borracha_psg_d":   request.form.get("borracha_psg_d", ""),
        "acc_borr_mtr":         request.form.get("borracha_mtr_d", ""),
        "acc_asa_urubu_dd":     request.form.get("asa_urubu_dd", ""),
        "acc_asa_urub_td":      request.form.get("asa_urubu_td", ""),
        "acc_tapete_mala":      request.form.get("tapete_mala", ""),
        "acc_tampa_prx":        request.form.get("tampa_paraxq", ""),
        "acc_borracha_psg_t":   request.form.get("borracha_psg_t", ""),
        "acc_borr_mtr_t":       request.form.get("borracha_mtr_t", ""),
        "acc_asa_urubu_de":     request.form.get("asa_urubu_de", ""),
        "acc_asa_urub_te":      request.form.get("asa_urubu_te", ""),
        "acc_bagagito":         request.form.get("bagagito", ""),
        "acc_linguet":          request.form.get("lingueta", ""),
    }

    # ── Salvar fotos temporariamente ──────────────────────────────────────────
    fotos_paths = []
    for foto in request.files.getlist("fotos"):
        if foto and foto.filename:
            safe = secure_filename(foto.filename)
            ext = Path(safe).suffix.lower()
            if ext in (".jpg", ".jpeg", ".png"):
                p = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
                foto.save(str(p))
                fotos_paths.append(p)

    formato = request.form.get("formato", "pdf").lower()

    # ── Gerar arquivo ─────────────────────────────────────────────────────────
    placa = _slugify(dados.get("placa", "PLACA"))
    data_slug = datetime.now().strftime("%d.%m.%Y")
    nome_docx = f"VISTORIA_{placa}_{data_slug}.docx"
    nome_pdf  = f"VISTORIA_{placa}_{data_slug}.pdf"
    caminho_docx = str(CONTRATOS_FOLDER / nome_docx)
    caminho_pdf  = str(CONTRATOS_FOLDER / nome_pdf)

    try:
        gerar_vistoria_entrega(dados, fotos_paths, caminho_docx, str(VISTORIA_TEMPLATE))
    except Exception as e:
        import traceback; traceback.print_exc()
        for p in fotos_paths:
            p.unlink(missing_ok=True)
        return jsonify({"error": f"Erro ao gerar vistoria: {e}"}), 500

    for p in fotos_paths:
        p.unlink(missing_ok=True)

    # ── Histórico + resposta ──────────────────────────────────────────────────
    if formato == "docx":
        _historico_append(dados.get("cliente", ""), "VISTORIA", nome_docx)
        return jsonify({"download_url": url_for("baixar_vistoria", nome=nome_docx)})

    try:
        _converter_pdf(caminho_docx, caminho_pdf)
    except Exception as e:
        flash(f"Erro ao converter para PDF: {e}", "erro")
        return redirect(url_for("pagina_vistoria"))

    _historico_append(dados.get("cliente", ""), "VISTORIA", nome_pdf)
    return jsonify({"download_url": url_for("baixar_vistoria", nome=nome_pdf)})


@app.route("/vistoria/gerar", methods=["POST"])
def gerar_vistoria_route():
    foto_path = None
    try:
        return _gerar_vistoria_impl()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro interno: {e}"}), 500
    finally:
        if foto_path:
            Path(foto_path).unlink(missing_ok=True)


def _gerar_vistoria_impl():
    agora = datetime.now(_BRT)
    dados = {
        "cliente_nome":      request.form.get("cliente_nome", ""),
        "cliente_telefone":  request.form.get("cliente_telefone", ""),
        "cliente_endereco":  request.form.get("cliente_endereco", ""),
        "preenchido_por":    request.form.get("preenchido_por", ""),
        "data_vistoria":     agora.strftime("%d/%m/%Y"),
        "veiculo":           request.form.get("veiculo", ""),
        "placa":             request.form.get("placa", "").upper(),
        "cor":               request.form.get("cor", ""),
        "ano":               request.form.get("ano", ""),
        "chassi":            request.form.get("chassi", ""),
        "numero_motor":      request.form.get("numero_motor", ""),
        "data_hora":         agora.strftime("%d/%m/%Y %H:%M"),
        "hodometro_entrega": request.form.get("hodometro_entrega", ""),
        "hodometro_retorno": request.form.get("hodometro_retorno", ""),
        "combustivel":       request.form.get("combustivel", ""),
        "acc_calotas":          request.form.get("acc_calotas", ""),
        "acc_buzina":           request.form.get("acc_buzina", ""),
        "acc_doc_crlv":         request.form.get("acc_doc_crlv", ""),
        "acc_triangulo":        request.form.get("acc_triangulo", ""),
        "acc_antena":           request.form.get("acc_antena", ""),
        "acc_sensor_re":        request.form.get("acc_sensor_re", ""),
        "acc_som":              request.form.get("acc_som", ""),
        "acc_tapetes":          request.form.get("acc_tapetes", ""),
        "acc_limpadores":       request.form.get("acc_limpadores", ""),
        "acc_chave_roda":       request.form.get("acc_chave_roda", ""),
        "acc_vidros_eletricos": request.form.get("acc_vidros_eletricos", ""),
        "acc_oleo_motor":       request.form.get("acc_oleo_motor", ""),
        "acc_alarme":           request.form.get("acc_alarme", ""),
        "acc_lampadas":         request.form.get("acc_lampadas", ""),
        "acc_macaco":           request.form.get("acc_macaco", ""),
        "acc_estepe":           request.form.get("acc_estepe", ""),
        "acc_gnv":              request.form.get("acc_gnv", ""),
        "acc_agua":             request.form.get("acc_agua", ""),
        "acc_borr_psg_dir":     request.form.get("acc_borr_psg_dir", ""),
        "acc_borr_mtr_dir":     request.form.get("acc_borr_mtr_dir", ""),
        "acc_asa_dd":           request.form.get("acc_asa_dd", ""),
        "acc_asa_td":           request.form.get("acc_asa_td", ""),
        "acc_tapete_mala":      request.form.get("acc_tapete_mala", ""),
        "acc_tampa_parachoque": request.form.get("acc_tampa_parachoque", ""),
        "acc_borr_psg_tras":    request.form.get("acc_borr_psg_tras", ""),
        "acc_borr_mtr_tras":    request.form.get("acc_borr_mtr_tras", ""),
        "acc_asa_de":           request.form.get("acc_asa_de", ""),
        "acc_asa_te":           request.form.get("acc_asa_te", ""),
        "acc_bagagito":         request.form.get("acc_bagagito", ""),
        "acc_lingueta":         request.form.get("acc_lingueta", ""),
        "obs_gerais":        request.form.get("obs_gerais", ""),
        "desc_sintomas":     request.form.get("desc_sintomas", ""),
    }

    foto_path = None
    foto = request.files.get("foto")
    if foto and foto.filename:
        ext = Path(secure_filename(foto.filename)).suffix.lower()
        if ext in ('.jpg', '.jpeg', '.png'):
            p = TEMP_FOLDER / f"{uuid.uuid4().hex}{ext}"
            foto.save(str(p))
            foto_path = str(p)

    edit_id    = request.form.get("edit_id", "").strip()
    placa_slug = _slugify(dados["placa"] or "PLACA")
    data_slug  = agora.strftime("%d.%m.%Y")
    nome_docx    = f"VISTORIA_{placa_slug}_{data_slug}.docx"
    caminho_docx = str(CONTRATOS_FOLDER / nome_docx)

    try:
        gerar_vistoria_nova(dados, foto_path, caminho_docx)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": f"Erro ao gerar vistoria: {e}"}), 500
    finally:
        if foto_path:
            Path(foto_path).unlink(missing_ok=True)

    # ── Supabase: INSERT na thread principal (rápido), upload no background ────
    if _os.environ.get("SUPABASE_URL") and _os.environ.get("SUPABASE_KEY"):
        import threading, traceback as _tb
        _storage_path = f"vistorias/{nome_docx}"
        _docx_bytes   = Path(caminho_docx).read_bytes()
        _insert       = {
            "cliente":           dados["cliente_nome"],
            "telefone":          dados["cliente_telefone"],
            "endereco":          dados["cliente_endereco"],
            "preenchido_por":    dados["preenchido_por"],
            "veiculo":           dados["veiculo"],
            "placa":             dados["placa"],
            "cor":               dados["cor"],
            "ano":               dados["ano"],
            "chassi":            dados["chassi"],
            "numero_motor":      dados["numero_motor"],
            "data_hora":         dados["data_hora"],
            "hodometro_entrega": dados["hodometro_entrega"],
            "hodometro_retorno": dados["hodometro_retorno"],
            "combustivel":       dados["combustivel"],
            "obs_gerais":        dados["obs_gerais"],
            "desc_sintomas":     dados["desc_sintomas"],
            "arquivo_path":      _storage_path,
            "acessorios":        {k: v for k, v in dados.items() if k.startswith('acc_')},
        }

        _old_storage_path = None
        sb = _supabase()
        if sb:
            try:
                sb.table("vistorias").insert(_insert).execute()
            except Exception:
                _tb.print_exc()
            if edit_id:
                try:
                    old = sb.table("vistorias").select("arquivo_path").eq("id", edit_id).single().execute()
                    _old_storage_path = (old.data or {}).get("arquivo_path")
                    sb.table("vistorias").delete().eq("id", edit_id).execute()
                except Exception:
                    _tb.print_exc()

        _ost = _old_storage_path

        def _bg():
            try:
                sb2 = _supabase()
                if not sb2:
                    return
                try:
                    sb2.storage.from_("documentos").upload(
                        _storage_path, _docx_bytes,
                        {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                         "upsert": "true"},
                    )
                except Exception:
                    _tb.print_exc()
                if _ost and _ost != _storage_path:
                    try:
                        sb2.storage.from_("documentos").remove([_ost])
                    except Exception:
                        pass
            except Exception:
                _tb.print_exc()

        threading.Thread(target=_bg, daemon=True).start()

    try:
        _historico_append(dados["cliente_nome"], "VISTORIA", nome_docx)
    except Exception:
        import traceback; traceback.print_exc()

    return jsonify({"redirect_url": url_for("historico_vistorias")})


# ── Histórico de Vistorias (Supabase) ─────────────────────────────────────────

@app.route("/historico/vistorias")
def historico_vistorias():
    sb = _supabase()
    vistorias = []
    erro = None
    if sb:
        try:
            res = sb.table("vistorias").select(
                "id, cliente, placa, veiculo, preenchido_por, data_hora, criado_em, arquivo_path"
            ).neq("deletado", True).order("criado_em", desc=True).execute()
            vistorias = res.data or []
        except Exception as e:
            erro = str(e)
    else:
        erro = "Supabase não configurado (SUPABASE_URL / SUPABASE_KEY ausentes)."
    return render_template("historico_vistorias.html", vistorias=vistorias, erro=erro, active="hist_vistorias")


@app.route("/historico/vistorias/<vistoria_id>/excluir", methods=["POST"])
def excluir_vistoria(vistoria_id):
    sb = _supabase()
    if sb:
        try:
            sb.table("vistorias").update({"deletado": True}).eq("id", vistoria_id).execute()
        except Exception:
            import traceback; traceback.print_exc()
    return jsonify({"ok": True})


@app.route("/historico/vistorias/download/<vistoria_id>")
def download_vistoria_supabase(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("arquivo_path").eq("id", vistoria_id).single().execute()
        path = res.data["arquivo_path"]
        signed = sb.storage.from_("documentos").create_signed_url(path, 60)
        return redirect(signed["signedURL"])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/historico/vistorias/<vistoria_id>/editar")
def editar_vistoria(vistoria_id):
    sb = _supabase()
    if not sb:
        flash("Supabase não configurado.", "erro")
        return redirect(url_for("historico_vistorias"))
    try:
        res = sb.table("vistorias").select("*").eq("id", vistoria_id).single().execute()
        vistoria = res.data
    except Exception as e:
        flash(f"Erro ao buscar vistoria: {e}", "erro")
        return redirect(url_for("historico_vistorias"))
    return render_template("vistoria.html", active="vistoria", vistoria=vistoria,
                           edit_id=vistoria_id, acessorios=vistoria.get("acessorios") or {})


@app.route("/historico/vistorias/download/<vistoria_id>/pdf")
def download_vistoria_pdf(vistoria_id):
    sb = _supabase()
    if not sb:
        abort(503)
    try:
        res = sb.table("vistorias").select("arquivo_path").eq("id", vistoria_id).single().execute()
        docx_storage_path = res.data["arquivo_path"]
        docx_bytes = sb.storage.from_("documentos").download(docx_storage_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    tmp_docx = TEMP_FOLDER / f"{uuid.uuid4().hex}.docx"
    tmp_pdf  = tmp_docx.with_suffix(".pdf")
    try:
        tmp_docx.write_bytes(docx_bytes)
        _converter_pdf(str(tmp_docx), str(tmp_pdf))
        pdf_bytes = tmp_pdf.read_bytes()
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar PDF: {e}"}), 422
    finally:
        tmp_docx.unlink(missing_ok=True)
        tmp_pdf.unlink(missing_ok=True)

    nome_pdf = Path(docx_storage_path).stem + ".pdf"
    return send_file(BytesIO(pdf_bytes), as_attachment=True, download_name=nome_pdf, mimetype="application/pdf")


@app.route("/vistoria/download/<nome>")
def baixar_vistoria(nome):
    caminho = CONTRATOS_FOLDER / nome
    if not caminho.exists() or caminho.parent.resolve() != CONTRATOS_FOLDER.resolve():
        abort(404)
    ext = caminho.suffix.lower()
    mime = "application/pdf" if ext == ".pdf" else \
           "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return send_file(str(caminho), as_attachment=True, download_name=nome, mimetype=mime)


if __name__ == "__main__":
    app.run(debug=True)
